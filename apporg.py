from flask import Flask, render_template, request, redirect, session, url_for, flash,send_file
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
import os
from datetime import datetime, timedelta
import jinja2
import pdfkit
from openpyxl import load_workbook
from flask import jsonify



data_folder = 'D:\\FOL_WEBAPP\\fuel_tracking' 
fuel_types = ['HSD BS-VI', 'Gas 87 MT', 'Oil 20W 50', 'Oil SG-240', 'Grease XG-279', 'Oil 80W 90']
date_format = "%d-%m-%Y"  # Date format for file naming
fuel_types = ["Fuel Type", "Oil Type", "Lubricant Type"]
app = Flask(__name__)
app.secret_key = 'your-secret-key'

# Configure SQLite database
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///D:/FOL_WEBAPP/instance/site.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
# User model
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(50), nullable=False)

# Manually create tables before running the app
def create_tables():
    with app.app_context():
        db.create_all()

# Run the table creation before app starts
create_tables()

EXCEL_FILE = 'D:\\FOL_WEBAPP\\fol_stock.xlsx'

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure the upload directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def get_file_path(vehicle_class):
    """Returns the appropriate file path based on vehicle class (A VEH or B VEH)."""
    veh=vehicle_class.lower()
    
    filename = load_weekly_excel(veh)

    return os.path.join(filename)

def update_fuel_log(data):
    """Updates the appropriate Excel file based on Vehicle Class and Vehicle Number."""
    vehicle_class = data["Vehicle Class"]
    file_path = get_file_path(vehicle_class)

    # Load existing data or create new DataFrame
    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
    else:
        # Create a new DataFrame with the required columns if the file doesn't exist
        df = pd.DataFrame(columns=[
            "Date", "Vehicle Class", "Vehicle Type", "Vehicle Number", "Fuel Type", 
            "Fuel Consumption", "Oil Type", "Oil Consumption", 
            "Lubricant Type", "Lubricant Consumption"
        ])

    # Convert date to string for comparison
    DATE_FORMAT = "%Y-%m-%d"
    data["Date"] = datetime.strptime(data["Date"], "%d-%m-%Y").strftime(DATE_FORMAT)

    # Check if the combination of Date and Vehicle Number already exists
    existing_entry = df[(df["Date"] == data["Date"]) & (df["Vehicle Number"] == data["Vehicle Number"])]

    if not existing_entry.empty:
        # Update the existing entry if a match is found
        df.loc[(df["Date"] == data["Date"]) & (df["Vehicle Number"] == data["Vehicle Number"]), df.columns] = data.values()
    else:
        # If no match is found, add a new row to the DataFrame
        df = pd.concat([df, pd.DataFrame([data])], ignore_index=True)

    # Save the updated DataFrame back to the Excel file
    df.to_excel(file_path, index=False)
# Function to load or create weekly Excel file
def load_or_create_excel(date):
    start_date = date - timedelta(days=date.weekday())  # Start of the week (Monday)
    end_date = start_date + timedelta(days=6)
    filename = os.path.join(data_folder, f'fuel_tracking_{start_date.strftime(date_format)}_to_{end_date.strftime(date_format)}.xlsx')

    print(f"Loading or creating Excel: {filename}")  # Debugging line

    if os.path.exists(filename):
        print(f"File exists: {filename}")
        df = pd.read_excel(filename)
    else:
        print(f"File does not exist, creating a new one")
        df = pd.DataFrame(columns=['Date', 'Order Number'] + fuel_types)
        df.to_excel(filename, index=False)

    return df, filename

# Function to create weekly Excel file
def create_weekly_excel(start_date, order_number):
    end_date = start_date + timedelta(days=6)
    filename = os.path.join(data_folder, f'fuel_tracking_{start_date.strftime(date_format)}_to_{end_date.strftime(date_format)}.xlsx')
    
    df = pd.DataFrame(columns=['Date', 'Order Number'] + fuel_types)
    df.to_excel(filename, index=False)
    return filename

# Function to load weekly Excel file
def load_weekly_excel(veh):
    date=date=datetime.now()
    start_date = date - timedelta(days=date.weekday())  # Start of the week (Monday)
    end_date = start_date + timedelta(days=6)
    
   
    if(veh=='alpha'):
        data_folder = 'D:\\FOL_WEBAPP\\fuel_tracking\\a veh' 
        filename = os.path.join(data_folder, f'fuel_tracking_{start_date.strftime(date_format)}_to_{end_date.strftime(date_format)}.xlsx')
    elif(veh=='beta'):
        data_folder='D:\\FOL_WEBAPP\\fuel_tracking\\b veh'
        filename = os.path.join(data_folder, f'fuel_tracking_{start_date.strftime(date_format)}_to_{end_date.strftime(date_format)}.xlsx')
    print(f"Loading or creating Excel: {filename}")  # Debugging line
    
    return filename


# Function to update fuel consumption
def update_fuel_consumption( date, fuel_consumption):
    filename, _ = load_weekly_excel(date)
    df = pd.read_excel(filename)
    
    date_str = date.strftime(date_format)
    if date_str in df['Date'].values:
        df.loc[df['Date'] == date_str, fuel_types] = fuel_consumption
    else:
        new_row = pd.DataFrame([[date_str] + fuel_consumption], columns=['Date'] + fuel_types)
        df = pd.concat([df, new_row], ignore_index=True)
    
    df.to_excel(filename, index=False)




# Function to check if today is Sunday
def is_last_day_of_week(date):
    return date.weekday() == 6  # 6 corresponds to Sunday in Python's weekday()


def load_weekly_reports():
    reports = {}

    for folder in ["A_VEH", "B_VEH"]:
        folder_path = os.path.join(data_folder, folder)
        
        if os.path.exists(folder_path):
            reports[folder] = []
            for filename in os.listdir(folder_path):
                if filename.endswith('.xlsx'):
                    file_path = os.path.join(folder_path, filename)
                    df = pd.read_excel(file_path)
                    reports[folder].append((filename, df.to_html(classes='table table-striped')))
    
    return reports

def load_specific_report(filename):
    filepath = os.path.join(data_folder, filename)
    if os.path.exists(filepath):
        df = pd.read_excel(filepath)
        return df.to_html(classes='table table-striped')
    else:
        return None
def generate_voucher_pdf(current_indent_no, last_indent_no, authority, last_received_dt, email_address, start_date, end_date, total_consumption):
    today_date = datetime.today()
    last_week_date = today_date - timedelta(days=7)
   

    # Constant
    demand_type = 'Normal'
    From = '12 MECH INF'
    To = 'Supply Depot ASC, Babina'
    formation = 'HQ 27 Armd Bde'
    fol_indent = 'ICV BMP-II/IIK- 32, CMT-03, ARV-02, AAT- 02, Motor Cycle-06, Maruti Gypsy - 14, TATA Safari - 02. Mahindra Scorpio - 01, 2.5 Ton TATA- 13, ALS-11, HMV 4x4 -06, HMV 8x8 with MHC-02, HMV 6x6 with MHC-05, Army Bus - 01, Water Bowser 5000 Ltrs-03, LRV-02, Kolos Tatra 8x8-01, Genr Set all type-07, Genr set (BFSR)-03, Chg set (BMP-2K)-16, Chg set 500W-01, Air compressor-01, ALS Med Amb-01, Integrated MsI Sml (IMS)-01'

    # Variables
    current_date = today_date.strftime("%d %b %Y")
    monthandyear = today_date.strftime("%b %Y")
    requier = today_date - timedelta(days=3)
    requier_date = requier.strftime("%d %b %Y")
    last_order_date = last_week_date.strftime("%d %b %Y")
    


    fuel_keys = [
        'HSD BS-VI', 'Gas 87 MT', '20W50', 
        'Oil SG-240', 'XG-279', '80W90'
    ]

    # Fetch values dynamically using dictionary comprehension
    fuel_consumption = {fuel: total_consumption.get(fuel, 0) for fuel in fuel_keys}

    # Extract values
    hsd_r = fuel_consumption['HSD BS-VI']
    gas87_r = fuel_consumption['Gas 87 MT']
    oil20w_r = fuel_consumption['20W50']
    oilsg_r = fuel_consumption['Oil SG-240']
    greasexg_r = fuel_consumption['XG-279']
    oil80w_r = fuel_consumption['80W90']


    context = {
        'demand_type': demand_type,
        'From': From,
        'To': To,
        'formation': formation,
        'current_indent_no': current_indent_no,
        'current_date': current_date,
        'last_indent_no': last_indent_no,
        'last_order_date': last_order_date,
        'authority': authority,
        'requier_date': requier_date,
        'fol_indent': fol_indent,
        'month_year': monthandyear,
        'total_consumption': total_consumption,
        'start_date': start_date.strftime("%d %b %Y"),
        'end_date': end_date.strftime("%d %b %Y"),
        'hsd_r': hsd_r,
        'gas87_r': gas87_r,
        'oil20w_r': oil20w_r,
        'oilsg_r': oilsg_r,
        'greasexg_r': greasexg_r,
        'oil80w_r': oil80w_r,
        'last_received_dt': last_received_dt,
        'email_address': email_address
    }

    template_loader = jinja2.FileSystemLoader('./templates')
    template_env = jinja2.Environment(loader=template_loader)
    template = template_env.get_template("voucher_temp.html")
    output_text = template.render(context)

    # Specify the path to your wkhtmltopdf executable
    # Example configuration for Windows
    config = pdfkit.configuration(wkhtmltopdf="C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe")   #change file path

    # Generate PDF from HTML content
    pdfkit.from_string(output_text, "FOL_voucher.pdf", configuration=config)
def get_dashboard_route():
    user_role = session.get("role", "staff")
    return {
        'admin': '/admin_dashboard',
        'manager': '/manager_dashboard',
        'staff': '/staff_dashboard'
    }.get(user_role, '/staff_dashboard')

def fol_used(fuel_type, fuel_consumption, oil_type, oil_consumption, lubricant_type, lubricant_consumption):
    """Subtracts the used quantities from stock (case insensitive)."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active  # Assuming the first sheet contains stock data
    except Exception as e:
        return f"Error reading Excel file: {e}"  

    # Convert inputs to lowercase for case-insensitive comparison
    fuel_type = fuel_type.lower()
    oil_type = oil_type.lower()
    lubricant_type = lubricant_type.lower()
    print(fuel_type)
    print(oil_type)
    print(lubricant_type)
    

    for row in ws.iter_rows(min_row=2, values_only=False):  # Iterate through stock rows
        item_name = str(row[1].value).lower()  # Convert 'Item Name' to lowercase

        if item_name == fuel_type:
            row[2].value = max(0, row[2].value - float(fuel_consumption))  # Subtract fuel
        elif item_name == oil_type:
            row[2].value = max(0, row[2].value - float(oil_consumption))  # Subtract oil
        elif item_name == lubricant_type:
            row[2].value = max(0, row[2].value - float(lubricant_consumption))  # Subtract lubricant

    # Save changes
    try:
        wb.save(EXCEL_FILE)
        wb.close()
    except Exception as e:
        return f"Error writing to Excel: {e}"


def total_fuel_consumption(date):
    start_date = date - timedelta(days=date.weekday())  # Get Monday of the week
    end_date = start_date + timedelta(days=6)  # Get Sunday of the week

    # Load weekly Excel file
    filename, _ = load_weekly_excel(start_date)  
    df = pd.read_excel(filename)

    # Convert Date column to datetime
    df['Date'] = pd.to_datetime(df['Date'], format=date_format)

    # Filter data for the given week
    mask = (df['Date'] >= start_date) & (df['Date'] <= end_date)
    weekly_data = df.loc[mask]

    # Initialize dictionary for fuel, oil, and lubricant consumption
    total_consumption = {}

    # Sum fuel consumption
    if "Fuel Type" in df.columns and "Fuel Consumption" in df.columns:
        fuel_data = weekly_data.groupby("Fuel Type")["Fuel Consumption"].sum()
        total_consumption.update(fuel_data.to_dict())

    # Sum oil consumption
    if "Oil Type" in df.columns and "Oil Consumption" in df.columns:
        oil_data = weekly_data.groupby("Oil Type")["Oil Consumption"].sum()
        total_consumption.update(oil_data.to_dict())

    # Sum lubricant consumption
    if "Lubricant Type" in df.columns and "Lubricant Consumption" in df.columns:
        lubricant_data = weekly_data.groupby("Lubricant Type")["Lubricant Consumption"].sum()
        total_consumption.update(lubricant_data.to_dict())

    return total_consumption

@app.route('/generate_voucher', methods=['GET', 'POST'])
def generate_voucher():
    dashboard_url = get_dashboard_route()
    if request.method == 'POST':
        # Get form data
        current_indent_no = request.form.get('current_indent_no')
        last_indent_no = request.form.get('last_indent_no')
        authority = request.form.get('authority')
        last_received_dt = request.form.get('last_received_dt')
        email_address = request.form.get('email_address')
        uploaded_file = request.files['report_file']

        if uploaded_file.filename:
            filepath = os.path.join(data_folder, uploaded_file.filename)
            uploaded_file.save(filepath)

            df = pd.read_excel(filepath)

            # Initialize dictionary for total consumption
            total_consumption = {}

            # Sum fuel consumption
            if "Fuel Type" in df.columns and "Fuel Consumption" in df.columns:
                fuel_data = df.groupby("Fuel Type")["Fuel Consumption"].sum()
                total_consumption.update(fuel_data.to_dict())

            # Sum oil consumption
            if "Oil Type" in df.columns and "Oil Consumption" in df.columns:
                oil_data = df.groupby("Oil Type")["Oil Consumption"].sum()
                total_consumption.update(oil_data.to_dict())

            # Sum lubricant consumption
            if "Lubricant Type" in df.columns and "Lubricant Consumption" in df.columns:
                lubricant_data = df.groupby("Lubricant Type")["Lubricant Consumption"].sum()
                total_consumption.update(lubricant_data.to_dict())

            try:
                # Extract date range from filename
                filename_parts = uploaded_file.filename.replace('.xlsx', '').split('_to_')
                start_date = datetime.strptime(filename_parts[0].split('_')[-1], date_format)
                end_date = datetime.strptime(filename_parts[1], date_format)
            except (IndexError, ValueError):
                # Default date range if filename format is incorrect
                date = datetime.now().date()
                start_date = date - timedelta(days=7)
                end_date = date
        else:
            # No file uploaded, calculate total consumption from records
            date = datetime.now().date()
            start_date = date - timedelta(days=7)
            end_date = date
            total_consumption = total_fuel_consumption(date)

        # Generate the PDF voucher
        generate_voucher_pdf(
            current_indent_no, last_indent_no, authority, last_received_dt, email_address, 
            start_date, end_date, total_consumption
        )


        pdf_file = r"D:\FOL_WEBAPP\FOL_voucher.pdf"

        return render_template('success.html', dashboard_url=dashboard_url)
    
    return render_template("voucher_details.html", dashboard_url=dashboard_url)

@app.route('/dashboard')
def index():
    return render_template('index.html')
@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']

        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password) and user.role == role:
            session['username'] = username
            session['role'] = role
            if role == 'admin':
                return redirect('/admin_dashboard')
            elif role == 'manager':
                return redirect('/manager_dashboard')
            elif role == 'staff':
                return redirect('/staff_dashboard')
        else:
            flash('Invalid credentials or role. Please try again.', 'error')
            return redirect(url_for('login'))

    return render_template('login.html')
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']

        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('User already exists. Choose another username.', 'warning')
            return redirect(url_for('register'))

        hashed_password = generate_password_hash(password)
        new_user = User(name=name, username=username, password=hashed_password, role=role)
        db.session.add(new_user)
        db.session.commit()
        flash('Registration successful. Please log in.', 'success')
        return redirect('/login')

    return render_template('register.html')
@app.route('/admin_dashboard')
def admin_dashboard():
    if session.get('role') != 'admin':
        return redirect('/login')
    return render_template('index.html')

@app.route('/manager_dashboard')
def manager_dashboard():
    if session.get('role') != 'manager':
        return redirect('/login')
    return render_template('dashboard_manager.html')

@app.route('/staff_dashboard')
def staff_dashboard():
    if session.get('role') != 'staff':
        return redirect('/login')
    return render_template('dashboard_staff.html')


@app.route('/upload_report', methods=['POST'])
def upload_report():
    if 'report_file' not in request.files:
        return "No file part"

    file = request.files['report_file']
    
    if file.filename == '':
        return "No selected file"

    if file:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)

        return redirect(url_for('view_report', filename=file.filename))

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

# Ensure the fuel_tracking folder exists
os.makedirs("fuel_tracking", exist_ok=True)

@app.route('/log_fuel', methods=['GET', 'POST'])
def log_fuel():
    dashboard_url = get_dashboard_route()
    
    if request.method == 'POST':
        # Get the form data, including vehicle number
        date = request.form.get('date', '')
        vehicle_class = request.form.get('vehicle_class', '')
        vehicle_type = request.form.get('vehicle_type', '')
        vehicle_number = request.form.get('vehicle_number', '')  # New field for vehicle number
        fuel_type = request.form.get('fuel_type', '')
        fuel_consumption = request.form.get('fuel_consumption', '0')
        oil_type = request.form.get('oil_type', '')
        oil_consumption = request.form.get('oil_consumption', '0')
        lubricant_type = request.form.get('lubricant_type', '')
        lubricant_consumption = request.form.get('lubricant_consumption', '0')

        # Format the date
        formatted_date = datetime.strptime(date, "%Y-%m-%d").strftime("%d-%m-%Y") if date else ''
        
        # Create a dictionary to hold the data
        data = {
            "Date": formatted_date,
            "Vehicle Class": vehicle_class,
            "Vehicle Type": vehicle_type,
            "Vehicle Number": vehicle_number,  # Add the vehicle number here
            "Fuel Type": fuel_type,
            "Fuel Consumption": fuel_consumption,
            "Oil Type": oil_type,
            "Oil Consumption": oil_consumption,
            "Lubricant Type": lubricant_type,
            "Lubricant Consumption": lubricant_consumption
        }
        
        # Update the fuel log with the new data
        update_fuel_log(data)
        
        # Call function to track fuel, oil, and lubricant usage
        fol_used(fuel_type, fuel_consumption, oil_type, oil_consumption, lubricant_type, lubricant_consumption)

        return render_template('success.html', dashboard_url=dashboard_url)
    
    return render_template("log_fol.html", dashboard_url=dashboard_url)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
@app.route('/view_reports')
def view_reports():
    files = [f for f in os.listdir(UPLOAD_FOLDER) if f.endswith('.xlsx') or f.endswith('.xls')]
    dashboard_url = get_dashboard_route()
    return render_template('view_report.html', reports=files, dashboard_url=dashboard_url)


@app.route('/view_report/<filename>')
def view_report(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if os.path.exists(filepath):
        df = pd.read_excel(filepath)

        # Create chart data
        chart_data = []

        if "Fuel Type" in df.columns and "Fuel Consumption" in df.columns:
            fuel_group = df.groupby("Fuel Type")["Fuel Consumption"].sum()
            for fuel, value in fuel_group.items():
                chart_data.append({"category": fuel, "value": value, "type": "Fuel"})

        if "Oil Type" in df.columns and "Oil Consumption" in df.columns:
            oil_group = df.groupby("Oil Type")["Oil Consumption"].sum()
            for oil, value in oil_group.items():
                chart_data.append({"category": oil, "value": value, "type": "Oil"})

        if "Lubricant Type" in df.columns and "Lubricant Consumption" in df.columns:
            lub_group = df.groupby("Lubricant Type")["Lubricant Consumption"].sum()
            for lub, value in lub_group.items():
                chart_data.append({"category": lub, "value": value, "type": "Lubricant"})

        # Convert full table
        table_data = df.to_dict(orient="records")
        table_columns = list(df.columns)

        dashboard_url = get_dashboard_route()

        return render_template(
            "view_specific_report.html",
            filename=filename,
            chart_data=chart_data,
            table_data=table_data,
            table_columns=table_columns,
            dashboard_url=dashboard_url
        )
    else:
        return "File not found", 404

@app.route('/success')
def success():
    dashboard_url = get_dashboard_route()
    return render_template('success.html', dashboard_url=dashboard_url)


@app.route('/review_voucher')
def review_voucher():
    path_to_pdf = "FOL_voucher.pdf"
    return send_file(path_to_pdf)




@app.route('/fol_stock', methods=['GET', 'POST'])
def stock():
    selected_category = request.values.get("category", "fuel")  # Default: fuel

    dashboard_url = get_dashboard_route() 

    try:
        stock_df = pd.read_excel(EXCEL_FILE)
        stock_df.columns = stock_df.columns.str.strip()
    except Exception as e:
        return f"Error reading Excel file: {e}"

    stock_df = stock_df[stock_df["Category"].str.lower() == selected_category.lower()]

    if request.method == "POST":
        updated_df = stock_df.copy()
        
        for idx, row in stock_df.iterrows():
            item_key = row['Item Name'].replace(" ", "_").lower()
            if item_key in request.form and request.form[item_key]:
                try:
                    add_qty = float(request.form[item_key])
                    updated_df.at[idx, 'Quantity'] += add_qty
                except ValueError:
                    pass
        
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            for idx, row in updated_df.iterrows():
                for excel_idx, excel_row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if excel_row[1] == row["Item Name"]:
                        ws.cell(row=excel_idx, column=3, value=row["Quantity"])
            wb.save(EXCEL_FILE)
            wb.close()
        except Exception as e:
            return f"Error writing to Excel: {e}"

        stock_df = updated_df

    return render_template("fol_stock.html", selected_category=selected_category, stock_data=stock_df, dashboard_url=dashboard_url)


if __name__ == '__main__':
    app.run(debug=True)
